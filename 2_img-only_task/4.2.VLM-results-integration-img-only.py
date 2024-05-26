import os
import json
import re
from openpyxl import Workbook, load_workbook


def extract_info_from_text(text):
    """
    Extracts specific fields from a JSON-like string using regex.
    
    Parameters:
    text (str): Text containing the JSON-like content.
    
    Returns:
    dict: A dictionary containing the extracted fields.
    """
    pattern = (r'\{[^{}]*"1_TypeOfMedicalImaging":\s*"([^"]*)"[^{}]*'
               r'"2_SpecificImagingSequence":\s*"([^"]*)"[^{}]*'
               r'"3_UseOfContrast":\s*"([^"]*)"[^{}]*'
               r'"4_ImagePlane":\s*"([^"]*)"[^{}]*'
               r'"5_PartOfTheBodyImaged":\s*"([^"]*)"[^{}]*\}')
    
    match = re.search(pattern, text)
    if match:
        return {
            '1_TypeOfMedicalImaging': match.group(1),
            '2_SpecificImagingSequence': match.group(2),
            '3_UseOfContrast': match.group(3),
            '4_ImagePlane': match.group(4),
            '5_PartOfTheBodyImaged': match.group(5)
        }
    else:
        return {
            '1_TypeOfMedicalImaging': "",
            '2_SpecificImagingSequence': "",
            '3_UseOfContrast': "",
            '4_ImagePlane': "",
            '5_PartOfTheBodyImaged': ""
        }


def process_files_in_folder(folder_path):
    """
    Processes text files in a given folder and saves the extracted information
    into an Excel file.
    
    Parameters:
    folder_path (str): Path to the folder containing text files.
    """
    # Ensure the directory exists
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Initialize Excel workbook
    wb = Workbook()
    ws = wb.active

    # Set header
    ws.append(['case_number', '1_Type_Of_Medical_Imaging', 
               '2_Specific_Imaging_Sequence', '3_Use_Of_Contrast', 
               '4_Image_Plane', '5_Part_Of_The_Body_Imaged'])

    for i in range(1, 273):
        file_path = os.path.join(folder_path, f'img_page{i}_0.png.txt')
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                extracted_data = extract_info_from_text(content)
                ws.append([i, 
                           extracted_data['1_TypeOfMedicalImaging'], 
                           extracted_data['2_SpecificImagingSequence'],
                           extracted_data['3_UseOfContrast'],
                           extracted_data['4_ImagePlane'],
                           extracted_data['5_PartOfTheBodyImaged']])
        except FileNotFoundError:
            # Handle file not found case
            ws.append([i, '', '', '', '', ''])
            print(f'File not found: {file_path}')
        except json.JSONDecodeError:
            # Handle JSON decode error case
            ws.append([i, '', '', '', '', ''])
            print(f'Invalid JSON format: {file_path}')

    # Save the Excel file
    excel_file_path = os.path.join(folder_path, 'sum.xlsx')
    wb.save(excel_file_path)
    print(f'Excel file saved: {excel_file_path}')


def combine_excel_files(folder_paths, combined_file_path):
    """
    Combines multiple Excel files into a single Excel file with separate sheets.
    
    Parameters:
    folder_paths (list): List of folder paths containing Excel files to combine.
    combined_file_path (str): Path to save the combined Excel file.
    """
    # Initialize new Excel workbook
    new_wb = Workbook()
    new_wb.remove(new_wb.active)  # Remove the default sheet

    for folder_path in folder_paths:
        sum_file_path = os.path.join(folder_path, 'sum.xlsx')
        try:
            # Load existing Excel file
            wb = load_workbook(sum_file_path)
            ws = wb.active

            # Set sheet name using the last part of the folder path
            sheet_name = folder_path.split('/')[-1]

            # Add new sheet to combined workbook and copy content
            new_ws = new_wb.create_sheet(title=sheet_name)
            for row in ws:
                new_ws.append([cell.value for cell in row])
        except FileNotFoundError:
            print(f'File not found: {sum_file_path}')

    # Save the combined Excel file
    new_wb.save(combined_file_path)
    print(f'Combined Excel file saved: {combined_file_path}')


def main():
    folder_paths = [
        'gpt4v_result/gpt4v_result_temp_1_try1',
        'gpt4o_result/gpt4o_result_temp_1_try1',
        'gemini_result/gemini_result_temp_1_try1',
        'gemini_flash_result/gemini_flash_result_temp_1_try1',
        'Claude_result/Claude_result_temp_1_try1',
    ]

    for folder_path in folder_paths:
        process_files_in_folder(folder_path)

    combined_excel_file_path = 'combined_sum.xlsx'
    combine_excel_files(folder_paths, combined_excel_file_path)


if __name__ == "__main__":
    main()
